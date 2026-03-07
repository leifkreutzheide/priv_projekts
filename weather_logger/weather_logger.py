"""
weather_logger.py
=================
Downloads today's weather data for Leipzig, Germany from the Open-Meteo API
and appends it as a new column in weather_data.xlsx (sheet: WeatherLog).

INSTALLATION (run once in VS Code terminal):
--------------------------------------------
pip install requests pandas openpyxl

RUN:
----
python weather_logger.py

HOW TO CHANGE LOCATION:
-----------------------
Edit LATITUDE and LONGITUDE below.
Find coordinates at: https://www.latlong.net/

HOW TO ADD MORE WEATHER VARIABLES:
-----------------------------------
1. Browse all available daily variables at:
   https://open-meteo.com/en/docs  (scroll to "Daily Weather Variables")
2. Copy the variable name (e.g. "uv_index_max")
3. Add it to WEATHER_VARIABLES below
4. Add a matching human-readable label to VARIABLE_LABELS in the same position

EXCEL LAYOUT AFTER SEVERAL RUNS:
----------------------------------
Row 1  : [Variable]         | Run 1      | Run 2      | Run 3      |  <- auto-numbered
Row 2  : Date               | 2025-06-01 | 2025-06-03 | 2025-06-07 |  <- date as data row
Row 3  : Temp High (C)      | 24.1       | 19.5       | 22.0       |
Row 4  : Temp Average (C)   | 17.3       | 14.2       | 16.8       |
Row 5  : Temp Low (C)       | 10.5       | 8.9        | 11.2       |
Row 6  : Sunrise            | 04:58      | 04:56      | 04:51      |
Row 7  : Sunset             | 21:22      | 21:24      | 21:29      |
Row 8  : Precipitation (mm) | 0.0        | 11.4       | 2.1        |
Row 9  : Wind Speed (km/h)  | 15.2       | 31.0       | 18.4       |
Row 10 : Label (1-10)       |            | 6          |            |  <- YOU fill in manually

NOTE ON DESIGN:
- Columns are numbered sequentially (Run 1, Run 2, ...) - no gaps ever
- Date is just another data row, useful for your overview but ignored by ML
- Skipped days leave no empty columns - the next run simply becomes the next column
- Label row is left blank by the script for you to fill in manually each day
"""

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os
import sys

# -----------------------------------------
# CONFIGURATION - edit these as needed
# -----------------------------------------

LATITUDE  = 51.3397   # Leipzig, Germany
LONGITUDE = 12.3731

EXCEL_FILE = "weather_data.xlsx"
SHEET_NAME = "WeatherLog"

# Open-Meteo API variable names (must match API exactly)
WEATHER_VARIABLES = [
    "temperature_2m_max",
    "temperature_2m_mean",
    "temperature_2m_min",
    "sunrise",
    "sunset",
    "precipitation_sum",
    "wind_speed_10m_max",
]

# Human-readable labels written into column A (must match order above)
VARIABLE_LABELS = [
    "Temp High (C)",
    "Temp Average (C)",
    "Temp Low (C)",
    "Sunrise",
    "Sunset",
    "Precipitation (mm)",
    "Wind Speed Max (km/h)",
]

LABEL_ROW_TEXT = "Label (1-10)"   # bottom row, filled in manually

# -----------------------------------------
# ROW INDEX CONSTANTS
# Row 1 : corner header
# Row 2 : Date
# Row 3 : first weather variable
# ...
# Last  : Label (1-10)
# -----------------------------------------

DATE_ROW    = 2
DATA_START  = 3                                      # first weather variable row
LABEL_ROW   = DATA_START + len(VARIABLE_LABELS)     # e.g. row 10

# -----------------------------------------
# STYLES
# -----------------------------------------

HEADER_FILL   = PatternFill("solid", start_color="2E4057")  # dark navy
RUN_FILL      = PatternFill("solid", start_color="048A81")  # teal  - run number header
DATE_FILL     = PatternFill("solid", start_color="E8F4F8")  # pale blue - date row
LABEL_FILL    = PatternFill("solid", start_color="F0F4F8")  # light grey - variable labels
ML_LABEL_FILL = PatternFill("solid", start_color="FFF3CD")  # amber - Label row

WHITE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DARK_FONT   = Font(name="Arial", bold=True, color="2E4057", size=10)
DATE_FONT   = Font(name="Arial", color="2E4057", size=10)
AMBER_FONT  = Font(name="Arial", bold=True, color="856404", size=10)
PLAIN_FONT  = Font(name="Arial", size=10)

CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# -----------------------------------------
# FETCH DATA
# -----------------------------------------

def fetch_weather():
    today = date.today().isoformat()
    url = "https://api.open-meteo.com/v1/forecast"
    params = {
        "latitude":        LATITUDE,
        "longitude":       LONGITUDE,
        "daily":           ",".join(WEATHER_VARIABLES),
        "start_date":      today,
        "end_date":        today,
        "timezone":        "Europe/Berlin",
        "wind_speed_unit": "kmh",
    }
    print(f"Fetching weather data for {today} ...")
    resp = requests.get(url, params=params, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    daily = data.get("daily", {})
    values = []
    for var in WEATHER_VARIABLES:
        raw = daily.get(var, [None])[0]
        if var in ("sunrise", "sunset") and raw:
            raw = raw[11:16]   # "2025-06-01T04:58" -> "04:58"
        values.append(raw)

    return today, values

# -----------------------------------------
# WRITE COLUMN A LABELS
# -----------------------------------------

def write_col_a(ws):
    """Write all row labels into column A. Skips cells that already have a value."""

    def set_cell(row, value, fill, font, alignment):
        cell = ws.cell(row=row, column=1)
        if cell.value is None:
            cell.value = value
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = THIN_BORDER

    # Corner header
    set_cell(1, "Variable", HEADER_FILL, WHITE_FONT, CENTER)

    # Date row label
    set_cell(DATE_ROW, "Date", DATE_FILL, DARK_FONT, LEFT)

    # Weather variable labels
    for i, label in enumerate(VARIABLE_LABELS):
        set_cell(DATA_START + i, label, LABEL_FILL, DARK_FONT, LEFT)

    # Label (1-10) row
    set_cell(LABEL_ROW, LABEL_ROW_TEXT, ML_LABEL_FILL, AMBER_FONT, LEFT)

# -----------------------------------------
# WRITE ONE DATA COLUMN
# -----------------------------------------

def write_data_column(ws, col, run_number, today, values):
    """Write run header + date + weather values + blank label cell."""

    # Run number header
    rc = ws.cell(row=1, column=col, value=f"Run {run_number}")
    rc.fill = RUN_FILL; rc.font = WHITE_FONT
    rc.alignment = CENTER; rc.border = THIN_BORDER

    # Date value
    dc = ws.cell(row=DATE_ROW, column=col, value=today)
    dc.fill = DATE_FILL; dc.font = DATE_FONT
    dc.alignment = CENTER; dc.border = THIN_BORDER

    # Weather values
    for i, val in enumerate(values):
        vc = ws.cell(row=DATA_START + i, column=col, value=val)
        vc.font = PLAIN_FONT; vc.alignment = CENTER; vc.border = THIN_BORDER

    # Label row - blank, for manual entry
    lc = ws.cell(row=LABEL_ROW, column=col, value=None)
    lc.fill = ML_LABEL_FILL; lc.border = THIN_BORDER; lc.alignment = CENTER

    ws.column_dimensions[get_column_letter(col)].width = 14

# -----------------------------------------
# CREATE OR APPEND
# -----------------------------------------

def build_new_workbook(today, values):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    write_col_a(ws)
    write_data_column(ws, col=2, run_number=1, today=today, values=values)

    ws.column_dimensions["A"].width = 22
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 18

    wb.save(EXCEL_FILE)
    print(f"Created '{EXCEL_FILE}' — Run 1 written for {today}.")


def append_to_existing(today, values):
    wb = load_workbook(EXCEL_FILE)

    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)

    ws = wb[SHEET_NAME]
    write_col_a(ws)

    # Find next empty column
    next_col = 2
    while ws.cell(row=1, column=next_col).value is not None:
        next_col += 1

    run_number = next_col - 1  # column 2 = Run 1, column 3 = Run 2, etc.

    # Guard against duplicate date (checks the Date row)
    for col in range(2, next_col):
        if ws.cell(row=DATE_ROW, column=col).value == today:
            print(f"Data for {today} already exists (Run {col - 1}). Skipping.")
            wb.close()
            return

    write_data_column(ws, col=next_col, run_number=run_number, today=today, values=values)

    ws.column_dimensions["A"].width = 22
    ws.freeze_panes = "B2"

    wb.save(EXCEL_FILE)
    print(f"Appended Run {run_number} for {today} to column {next_col} in '{EXCEL_FILE}'.")

# -----------------------------------------
# MAIN
# -----------------------------------------

if __name__ == "__main__":
    try:
        today, values = fetch_weather()
    except requests.exceptions.RequestException as e:
        print(f"Network error: {e}")
        sys.exit(1)

    if not os.path.exists(EXCEL_FILE):
        build_new_workbook(today, values)
    else:
        append_to_existing(today, values)

    print("Done.")
